from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import Account
from app.services.collector import Collector
from app.services.exporter import WorkbookExporter
from app.services.token_vault import TokenVault


@pytest.mark.asyncio
async def test_export_contains_formula_driven_report(
    database,
    settings,
    fake_client_factory,
):
    _engine, session_factory = database
    vault = TokenVault(settings.data_dir)
    with session_factory() as session:
        account = Account(
            name="Тестовый кабинет",
            encrypted_token=vault.encrypt("valid-token"),
        )
        session.add(account)
        session.commit()
        account_id = account.id

    target_date = date(2026, 7, 27)
    collector = Collector(session_factory, vault, fake_client_factory)
    await collector.collect(
        account_id,
        target_date,
        target_date,
        trigger="manual",
    )

    payload, filename = WorkbookExporter(session_factory).build(
        account_id,
        target_date,
        target_date,
    )
    workbook = load_workbook(BytesIO(payload), data_only=False)

    assert filename.endswith(".xlsx")
    assert "Целевые показатели по РК" in workbook.sheetnames
    assert workbook.sheetnames[:2] == [
        "Итого по группам",
        "По артикулам",
    ]
    assert workbook["Исходные данные"].sheet_state == "hidden"

    summary = workbook["Итого по группам"]
    assert summary["A3"].value == "Фотообои"
    assert summary["B3"].value == "Показы"
    assert summary["D3"].value.startswith("=SUMIFS(")
    assert "D$1" in summary["D3"].value
    assert summary["D2"].value == "27.07"
    assert summary["D4"].value == "=IFERROR(D8/D3*1000,0)"
    assert summary["D6"].value == "=IFERROR(D5/D3,0)"
    assert summary["D15"].value == "=IFERROR(D8/D13,0)"

    for row in range(3, 16):
        assert str(summary.cell(row, 3).value).startswith("=")
        assert str(summary.cell(row, 4).value).startswith("=")

    detail = workbook["По артикулам"]
    assert detail["A3"].value == "Фотообои"
    assert detail["B3"].value == "Фотообои Горы"
    assert detail["C3"].value == 699712395
    assert detail["D3"].value == "Показы"
    assert detail["F3"].value.startswith("=SUMIFS(")
    assert "F$1" in detail["F3"].value
    assert detail.row_dimensions[4].outline_level == 1
    assert detail.row_dimensions[4].hidden is False

    targets = workbook["Целевые показатели по РК"]
    assert targets["A1"].value == "Метрика ФОТООБОИ"
    assert targets["D13"].value == "<10%"
