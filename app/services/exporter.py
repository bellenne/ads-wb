from __future__ import annotations

from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, sessionmaker

from app.models import Account, Campaign, DailyStat, Product
from app.services.grouping import canonical_report_group

BLUE = "6FA8DC"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
RED = "F4CCCC"
YELLOW = "FFF2CC"
GREEN = "D9EAD3"
WHITE = "FFFFFF"
TEXT = "1F2933"
GRID = "D8E1E8"
GROUP_FILL = "CFE8F3"

METRICS = [
    ("views", "Показы", "count"),
    ("cpm", "CPM, ₽ (ставка)", "currency"),
    ("clicks", "Клики", "count"),
    ("ctr", "CTR, %", "percent"),
    ("cpc", "CPC, ₽ (цена клика)", "currency"),
    ("spend", "Затраты", "currency"),
    ("atbs", "Корзина", "count"),
    ("cart_cr", "Конверсия в корзину, %", "percent"),
    ("orders", "Заказы", "count"),
    ("order_cr", "Конверсия в заказ, %", "percent"),
    ("revenue", "Заказали на сумму, ₽", "currency"),
    ("cpo", "CPO, ₽ (цена заказа)", "currency"),
    ("drr", "ДРРЗ, %", "percent"),
]

RAW_COLUMNS = {
    "date": "A",
    "group": "B",
    "nm_id": "C",
    "product": "D",
    "advert_id": "E",
    "campaign": "F",
    "views": "G",
    "clicks": "H",
    "spend": "I",
    "atbs": "J",
    "orders": "K",
    "shks": "L",
    "revenue": "M",
    "canceled": "N",
}


class WorkbookExporter:
    def __init__(self, session_factory: sessionmaker[Session]) -> None:
        self.session_factory = session_factory

    def build(
        self,
        account_id: int,
        begin: date,
        end: date,
    ) -> tuple[bytes, str]:
        if begin > end:
            raise ValueError("Дата начала не может быть позже даты окончания")
        account, rows = self._load_rows(account_id, begin, end)
        if account is None:
            raise ValueError("Кабинет не найден")

        workbook = Workbook()
        workbook.remove(workbook.active)
        group_sheet = workbook.create_sheet("Итого по группам")
        product_sheet = workbook.create_sheet("По артикулам")
        target_sheet = workbook.create_sheet("Целевые показатели по РК")
        raw_sheet = workbook.create_sheet("Исходные данные")

        self._write_targets(target_sheet)
        raw_last_row = self._write_raw_data(raw_sheet, rows)
        self._write_group_report(
            group_sheet,
            rows,
            begin,
            end,
            raw_last_row,
            account.name,
        )
        self._write_product_report(
            product_sheet,
            rows,
            begin,
            end,
            raw_last_row,
            account.name,
        )
        raw_sheet.sheet_state = "hidden"

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

        stream = BytesIO()
        workbook.save(stream)
        filename = (
            f"Статистика РК {begin.isoformat()}—{end.isoformat()}.xlsx"
        )
        return stream.getvalue(), filename

    def _load_rows(
        self,
        account_id: int,
        begin: date,
        end: date,
    ) -> tuple[Account | None, list[dict[str, Any]]]:
        with self.session_factory() as session:
            account = session.get(Account, account_id)
            statement = (
                select(DailyStat, Product, Campaign)
                .outerjoin(
                    Product,
                    (Product.account_id == DailyStat.account_id)
                    & (Product.nm_id == DailyStat.nm_id),
                )
                .outerjoin(
                    Campaign,
                    (Campaign.account_id == DailyStat.account_id)
                    & (Campaign.advert_id == DailyStat.advert_id),
                )
                .where(
                    DailyStat.account_id == account_id,
                    DailyStat.stat_date >= begin,
                    DailyStat.stat_date <= end,
                )
                .order_by(
                    DailyStat.stat_date,
                    DailyStat.nm_id,
                    DailyStat.advert_id,
                )
            )
            rows = []
            for stat, product, campaign in session.execute(statement):
                rows.append(
                    {
                        "date": stat.stat_date,
                        "group": canonical_report_group(
                            product.report_group if product else "",
                            product.subject_name if product else "",
                            product.name if product else "",
                            campaign.name if campaign else "",
                            is_manual=bool(
                                product and product.group_is_manual
                            ),
                        ),
                        "nm_id": stat.nm_id,
                        "product": product.name if product else "",
                        "advert_id": stat.advert_id,
                        "campaign": campaign.name if campaign else "",
                        "views": stat.views,
                        "clicks": stat.clicks,
                        "spend": Decimal(stat.spend),
                        "atbs": stat.atbs,
                        "orders": stat.orders,
                        "shks": stat.shks,
                        "revenue": Decimal(stat.revenue),
                        "canceled": stat.canceled,
                    }
                )
            return account, rows

    def _write_targets(self, sheet) -> None:
        photo_rows = [
            ["Метрика ФОТООБОИ", "Красная зона", "Желтая зона", "Зеленая зона"],
            ["CTR, %", "<2,5%", "2,5%-3,5%", ">3,5%"],
            ["CPC, ₽ (цена клика)", ">20 руб", "13-19 руб", "<13 руб"],
            ["Конверсия в корзину, %", "<6%", "7-9%", "10%>"],
            ["Конверсия в заказ%", "<5%", "6-8%", "9%>"],
            ["ДРРЗ", ">15%", "10%-15%", "<10%"],
        ]
        shirt_rows = [
            ["Метрика Футболки.", "Красная зона", "Желтая зона", "Зеленая зона"],
            ["CTR, %", "<2,5%", "2,5%-4,5%", ">4,5%"],
            ["CPC, ₽ (цена клика)", ">16 руб", "11-15 руб", "<10 руб"],
            ["Конверсия в корзину, %", "<6%", "7-10%", "11%>"],
            ["Конверсия в заказ%", "<20%", "21-26%", "27%>"],
            ["ДРРЗ", ">15%", "10%-15%", "<10%"],
        ]
        for row_index, values in enumerate(photo_rows, start=1):
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row_index, column_index, value)
        for row_index, values in enumerate(shirt_rows, start=8):
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row_index, column_index, value)

        for row_index in list(range(1, 7)) + list(range(8, 14)):
            sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            for column, color in ((2, RED), (3, YELLOW), (4, GREEN)):
                sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor=color
                )
            for cell in sheet[row_index][:4]:
                cell.font = Font(name="Montserrat", bold=True, color=TEXT)
                cell.alignment = Alignment(
                    horizontal="center" if cell.column > 1 else "left",
                    vertical="center",
                )

        for header_row in (1, 8):
            sheet.cell(header_row, 1).fill = PatternFill(
                "solid", fgColor=BLUE
            )
            sheet.cell(header_row, 1).font = Font(
                name="Montserrat", bold=True, color=TEXT
            )

        sheet["A15"] = (
            "Источник: https://dev.wildberries.ru/openapi/promotion — "
            "GET /adv/v3/fullstats"
        )
        sheet["A15"].font = Font(
            name="Montserrat", size=9, color="52616B", italic=True
        )
        sheet.merge_cells("A15:D15")
        sheet.column_dimensions["A"].width = 30
        for column in ("B", "C", "D"):
            sheet.column_dimensions[column].width = 18
        sheet.row_dimensions[15].height = 30
        sheet["A15"].alignment = Alignment(wrap_text=True, vertical="center")
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A1"

    def _write_raw_data(
        self,
        sheet,
        rows: list[dict[str, Any]],
    ) -> int:
        headers = [
            "Дата",
            "Группа",
            "Артикул WB",
            "Название товара",
            "ID кампании",
            "Название кампании",
            "Показы",
            "Клики",
            "Затраты",
            "Корзина",
            "Заказы",
            "Заказанные товары",
            "Заказали на сумму",
            "Отмены",
        ]
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    row["date"],
                    row["group"],
                    row["nm_id"],
                    row["product"],
                    row["advert_id"],
                    row["campaign"],
                    row["views"],
                    row["clicks"],
                    float(row["spend"]),
                    row["atbs"],
                    row["orders"],
                    row["shks"],
                    float(row["revenue"]),
                    row["canceled"],
                ]
            )

        header_fill = PatternFill("solid", fgColor=BLUE)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(name="Montserrat", bold=True, color=WHITE)
        for cell in sheet["A"][1:]:
            cell.number_format = "yyyy-mm-dd"
        for column in ("I", "M"):
            for cell in sheet[column][1:]:
                cell.number_format = '#,##0.00" ₽"'
        widths = {
            "A": 13,
            "B": 20,
            "C": 16,
            "D": 30,
            "E": 16,
            "F": 32,
            "G": 13,
            "H": 12,
            "I": 15,
            "J": 12,
            "K": 12,
            "L": 19,
            "M": 20,
            "N": 10,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[1].height = 30
        for cell in sheet[1]:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:N{max(sheet.max_row, 1)}"
        return max(sheet.max_row, 2)

    def _write_group_report(
        self,
        sheet,
        rows: list[dict[str, Any]],
        begin: date,
        end: date,
        raw_last_row: int,
        account_name: str,
    ) -> None:
        dates = self._dates(begin, end)
        groups = sorted(
            {str(row["group"]) for row in rows},
            key=str.casefold,
        )
        date_start_column = 4
        last_column = date_start_column - 1 + len(dates)
        self._write_report_header(
            sheet,
            account_name,
            ["Группа", "Показатель", "Итого"],
            dates,
            date_start_column,
        )
        current_row = 3
        for group in groups:
            current_row = self._write_metric_block(
                sheet,
                current_row,
                scope="group",
                labels=[group],
                criteria_value=group,
                criteria_cell_column=1,
                group_name=group,
                dates=dates,
                raw_last_row=raw_last_row,
                last_column=last_column,
                metric_column=2,
                total_column=3,
                date_start_column=date_start_column,
            )
            current_row += 1
        if not groups:
            self._write_empty_state(sheet, last_column)

        sheet.freeze_panes = "D3"
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 31
        sheet.column_dimensions["C"].width = 15
        for column in range(date_start_column, last_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 12
        sheet.auto_filter.ref = f"A2:{get_column_letter(last_column)}2"

    def _write_product_report(
        self,
        sheet,
        rows: list[dict[str, Any]],
        begin: date,
        end: date,
        raw_last_row: int,
        account_name: str,
    ) -> None:
        dates = self._dates(begin, end)
        grouped_products: defaultdict[str, dict[int, str]] = defaultdict(dict)
        for row in rows:
            grouped_products[row["group"]][row["nm_id"]] = (
                row["product"] or f"Артикул {row['nm_id']}"
            )

        date_start_column = 6
        last_column = date_start_column - 1 + len(dates)
        self._write_report_header(
            sheet,
            account_name,
            [
                "Группа",
                "Товар",
                "Артикул WB",
                "Показатель",
                "Итого",
            ],
            dates,
            date_start_column,
        )
        current_row = 3
        for group in sorted(grouped_products, key=str.casefold):
            for nm_id, product_name in sorted(
                grouped_products[group].items(),
                key=lambda item: (item[1].casefold(), item[0]),
            ):
                block_start = current_row
                current_row = self._write_metric_block(
                    sheet,
                    current_row,
                    scope="product",
                    labels=[group, product_name, nm_id],
                    criteria_value=nm_id,
                    criteria_cell_column=3,
                    group_name=group,
                    dates=dates,
                    raw_last_row=raw_last_row,
                    last_column=last_column,
                    metric_column=4,
                    total_column=5,
                    date_start_column=date_start_column,
                )
                sheet.row_dimensions.group(
                    block_start + 1,
                    current_row - 1,
                    outline_level=1,
                    hidden=False,
                )
                current_row += 1

        if not grouped_products:
            self._write_empty_state(sheet, last_column)

        sheet.sheet_properties.outlinePr.summaryBelow = False
        sheet.freeze_panes = "F3"
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 50
        sheet.column_dimensions["C"].width = 16
        sheet.column_dimensions["D"].width = 31
        sheet.column_dimensions["E"].width = 15
        for column in range(date_start_column, last_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 12
        sheet.auto_filter.ref = f"A2:{get_column_letter(last_column)}2"

    @staticmethod
    def _dates(begin: date, end: date) -> list[date]:
        dates = []
        cursor = begin
        while cursor <= end:
            dates.append(cursor)
            cursor += timedelta(days=1)
        return dates

    @staticmethod
    def _write_report_header(
        sheet,
        account_name: str,
        headers: list[str],
        dates: list[date],
        date_start_column: int,
    ) -> None:
        sheet["A1"] = f"Кабинет: {account_name}"
        sheet["A1"].font = Font(
            name="Montserrat", size=10, italic=True, color="52616B"
        )
        for column, header in enumerate(headers, start=1):
            sheet.cell(2, column, header)
        for offset, stat_date in enumerate(
            dates,
            start=date_start_column,
        ):
            helper = sheet.cell(1, offset, stat_date)
            helper.number_format = "yyyy-mm-dd"
            helper.font = Font(name="Montserrat", size=1, color=WHITE)
            sheet.cell(2, offset, stat_date.strftime("%d.%m"))

        last_column = date_start_column - 1 + len(dates)
        for cell in sheet[2][:last_column]:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(name="Montserrat", bold=True, color=WHITE)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        sheet.row_dimensions[2].height = 30

    @staticmethod
    def _write_empty_state(sheet, last_column: int) -> None:
        sheet["A4"] = (
            "За выбранный период данных нет. Запустите синхронизацию "
            "или измените период."
        )
        sheet.merge_cells(
            start_row=4,
            start_column=1,
            end_row=5,
            end_column=max(last_column, 8),
        )
        sheet["A4"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        sheet["A4"].font = Font(
            name="Montserrat", color="52616B", italic=True
        )

    def _write_metric_block(
        self,
        sheet,
        start_row: int,
        *,
        scope: str,
        labels: list[str | int],
        criteria_value: str | int,
        criteria_cell_column: int,
        group_name: str,
        dates: list[date],
        raw_last_row: int,
        last_column: int,
        metric_column: int,
        total_column: int,
        date_start_column: int,
    ) -> int:
        metric_rows: dict[str, int] = {}
        thin_bottom = Side(style="thin", color=GRID)

        for index, (key, label, number_kind) in enumerate(METRICS):
            row = start_row + index
            metric_rows[key] = row
            if index == 0:
                for column, value in enumerate(labels, start=1):
                    sheet.cell(row, column, value)
                sheet.cell(
                    row,
                    criteria_cell_column,
                    criteria_value,
                )
            sheet.cell(row, metric_column, label)

            for column in range(1, last_column + 1):
                cell = sheet.cell(row, column)
                cell.font = Font(name="Montserrat", size=9, color=TEXT)
                cell.border = Border(bottom=thin_bottom)
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal=(
                        "left" if column < total_column else "right"
                    ),
                )
                if number_kind == "count" and column >= total_column:
                    cell.number_format = "#,##0"
                elif number_kind == "currency" and column >= total_column:
                    cell.number_format = '#,##0.00" ₽"'
                elif number_kind == "percent" and column >= total_column:
                    cell.number_format = "0.00%"

        label_fill = PatternFill(
            "solid", fgColor=GROUP_FILL if scope == "group" else PALE_BLUE
        )
        for row in range(start_row, start_row + len(METRICS)):
            sheet.cell(row, metric_column).fill = label_fill
        for column in range(1, metric_column):
            sheet.cell(start_row, column).fill = label_fill
            sheet.cell(start_row, column).font = Font(
                name="Montserrat",
                bold=(scope == "group" or column == criteria_cell_column),
                color=TEXT,
            )

        criteria_column = RAW_COLUMNS["group" if scope == "group" else "nm_id"]
        criteria_cell_letter = get_column_letter(criteria_cell_column)
        criteria_cell = f"${criteria_cell_letter}${start_row}"
        raw_ranges = {
            key: (
                f"'Исходные данные'!${column}$2:${column}${raw_last_row}"
            )
            for key, column in RAW_COLUMNS.items()
        }
        criteria_range = (
            f"'Исходные данные'!${criteria_column}$2:"
            f"${criteria_column}${raw_last_row}"
        )
        date_range = raw_ranges["date"]

        for day_index, _stat_date in enumerate(
            dates,
            start=date_start_column,
        ):
            column_letter = get_column_letter(day_index)
            date_cell = f"{column_letter}$1"
            formulas = self._metric_formulas(
                metric_rows,
                column_letter,
                date_cell,
                criteria_range,
                criteria_cell,
                date_range,
                raw_ranges,
            )
            for key, formula in formulas.items():
                sheet.cell(metric_rows[key], day_index, formula)

        first_date_column = get_column_letter(date_start_column)
        last_date_column = get_column_letter(last_column)
        total_formulas = self._total_formulas(
            metric_rows,
            first_date_column,
            last_date_column,
            get_column_letter(total_column),
        )
        for key, formula in total_formulas.items():
            sheet.cell(metric_rows[key], total_column, formula)

        formula_note = (
            "Исходные показатели суммируются из скрытого листа "
            "«Исходные данные» через SUMIFS. Производные KPI считаются "
            "формулами из исходных показателей."
        )
        sheet.cell(start_row, metric_column).comment = Comment(
            formula_note,
            "User",
        )
        self._add_zone_formatting(
            sheet,
            metric_rows,
            total_column,
            last_column,
            is_shirt="футбол" in group_name.casefold(),
        )
        return start_row + len(METRICS)

    @staticmethod
    def _metric_formulas(
        rows: dict[str, int],
        column_letter: str,
        date_cell: str,
        criteria_range: str,
        criteria_cell: str,
        date_range: str,
        raw_ranges: dict[str, str],
    ) -> dict[str, str]:
        def sumifs(metric: str) -> str:
            return (
                f"=SUMIFS({raw_ranges[metric]},{date_range},{date_cell},"
                f"{criteria_range},{criteria_cell})"
            )

        def cell(key: str) -> str:
            return f"{column_letter}{rows[key]}"

        return {
            "views": sumifs("views"),
            "cpm": (
                f"=IFERROR({cell('spend')}/{cell('views')}*1000,0)"
            ),
            "clicks": sumifs("clicks"),
            "ctr": f"=IFERROR({cell('clicks')}/{cell('views')},0)",
            "cpc": f"=IFERROR({cell('spend')}/{cell('clicks')},0)",
            "spend": sumifs("spend"),
            "atbs": sumifs("atbs"),
            "cart_cr": f"=IFERROR({cell('atbs')}/{cell('clicks')},0)",
            "orders": sumifs("orders"),
            "order_cr": f"=IFERROR({cell('orders')}/{cell('atbs')},0)",
            "revenue": sumifs("revenue"),
            "cpo": f"=IFERROR({cell('spend')}/{cell('orders')},0)",
            "drr": f"=IFERROR({cell('spend')}/{cell('revenue')},0)",
        }

    @staticmethod
    def _total_formulas(
        rows: dict[str, int],
        first_column: str,
        last_column: str,
        total_column: str,
    ) -> dict[str, str]:
        def total(key: str) -> str:
            row = rows[key]
            return f"=SUM({first_column}{row}:{last_column}{row})"

        def cell(key: str) -> str:
            return f"{total_column}{rows[key]}"

        return {
            "views": total("views"),
            "cpm": (
                f"=IFERROR({cell('spend')}/{cell('views')}*1000,0)"
            ),
            "clicks": total("clicks"),
            "ctr": f"=IFERROR({cell('clicks')}/{cell('views')},0)",
            "cpc": f"=IFERROR({cell('spend')}/{cell('clicks')},0)",
            "spend": total("spend"),
            "atbs": total("atbs"),
            "cart_cr": f"=IFERROR({cell('atbs')}/{cell('clicks')},0)",
            "orders": total("orders"),
            "order_cr": f"=IFERROR({cell('orders')}/{cell('atbs')},0)",
            "revenue": total("revenue"),
            "cpo": f"=IFERROR({cell('spend')}/{cell('orders')},0)",
            "drr": f"=IFERROR({cell('spend')}/{cell('revenue')},0)",
        }

    @staticmethod
    def _add_zone_formatting(
        sheet,
        rows: dict[str, int],
        first_column: int,
        last_column: int,
        *,
        is_shirt: bool,
    ) -> None:
        fills = {
            "red": PatternFill("solid", fgColor=RED),
            "yellow": PatternFill("solid", fgColor=YELLOW),
            "green": PatternFill("solid", fgColor=GREEN),
        }
        rules = {
            "ctr": (
                0.025,
                0.045 if is_shirt else 0.035,
                "higher",
            ),
            "cpc": (
                11 if is_shirt else 13,
                16 if is_shirt else 20,
                "lower",
            ),
            "cart_cr": (
                0.07,
                0.11 if is_shirt else 0.10,
                "higher",
            ),
            "order_cr": (
                0.21 if is_shirt else 0.06,
                0.27 if is_shirt else 0.09,
                "higher",
            ),
            "drr": (0.10, 0.15, "lower"),
        }
        start_letter = get_column_letter(first_column)
        end_letter = get_column_letter(last_column)
        for key, (lower, upper, direction) in rules.items():
            row = rows[key]
            cell_range = f"{start_letter}{row}:{end_letter}{row}"
            if direction == "higher":
                sheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="lessThan",
                        formula=[str(lower)],
                        fill=fills["red"],
                        stopIfTrue=True,
                    ),
                )
                sheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="between",
                        formula=[str(lower), str(upper)],
                        fill=fills["yellow"],
                        stopIfTrue=True,
                    ),
                )
                sheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=[str(upper)],
                        fill=fills["green"],
                        stopIfTrue=True,
                    ),
                )
            else:
                sheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="lessThan",
                        formula=[str(lower)],
                        fill=fills["green"],
                        stopIfTrue=True,
                    ),
                )
                sheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="between",
                        formula=[str(lower), str(upper)],
                        fill=fills["yellow"],
                        stopIfTrue=True,
                    ),
                )
                sheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=[str(upper)],
                        fill=fills["red"],
                        stopIfTrue=True,
                    ),
                )
